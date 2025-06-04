from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.conf import settings
from openpyxl import load_workbook, Workbook
import os
from datetime import date, datetime
from .forms import FeedbackForm

# Directory setup
APP_DIR = os.path.dirname(os.path.abspath(__file__))
FEEDBACK_DIR = os.path.join(APP_DIR, "feedback_data")
os.makedirs(FEEDBACK_DIR, exist_ok=True)

def get_excel_file(state):
    return os.path.join(FEEDBACK_DIR, f"feedback_{state}.xlsx")

def get_student_session_number(student_name, selected_date, state):
    """Get the next session number for a student on a specific date"""
    file_path = get_excel_file(state.lower())
    if not os.path.exists(file_path):
        return 1
    
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        session_count = 0
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if (row and len(row) > 1 and row[0] and row[1] and 
                str(row[0]) == selected_date and 
                row[1].strip().lower() == student_name.strip().lower()):
                session_count += 1
        
        workbook.close()
        return session_count + 1
    except Exception as e:
        print(f"Error getting session number: {e}")
        return 1

def check_existing_incomplete_session(student_name, selected_date, state):
    """Check if there's an existing incomplete session (no login time) for the student"""
    file_path = get_excel_file(state.lower())
    if not os.path.exists(file_path):
        return None
    
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if (row and len(row) > 11 and row[0] and row[1] and 
                str(row[0]) == selected_date and 
                row[1].strip().lower() == student_name.strip().lower() and
                (not row[11] or row[11] == "")):  # No login time (column L is empty - index 11)
                workbook.close()
                session_num = row[10] if len(row) > 10 and row[10] else 1  # Session Number is column K (index 10)
                return row_num, session_num  # Return row number and session number
        
        workbook.close()
        return None
    except Exception as e:
        print(f"Error checking existing session: {e}")
        return None

def update_existing_session_with_login(student_name, selected_date, state, login_time, row_num):
    """Update existing session row with login time"""
    file_path = get_excel_file(state.lower())
    
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        
        # Update login time in the existing row (column L = index 12)
        sheet.cell(row=row_num, column=12, value=login_time)  # Login Time (column L)
        
        workbook.save(file_path)
        workbook.close()
        print(f"Updated existing session at row {row_num} with login time")
        return True
        
    except Exception as e:
        print(f"Error updating session with login: {e}")
        return False

def save_login_to_excel(student_name, selected_date, state, login_time, session_number):
    """Save student login data immediately to Excel - creates a new session row"""
    file_path = get_excel_file(state.lower())
    print(f"Saving login data to file: {file_path}")
    
    # Create file if it doesn't exist
    if not os.path.exists(file_path):
        print("Creating new Excel file for session data")
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Student Sessions"
        sheet.append([
            "Date", "Student Name", "Trainer Name", "Course", "Slot Timings", 
            "Understanding", "Engagement", "Overall Feedback", "Homework", 
            "Parents Feedback", "Session Number", "Login Time", "Logout Time"
        ])
        workbook.save(file_path)
        workbook.close()

    # Open existing file and append login data
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        
        # Create new session row with login data only
        login_row = [
            selected_date, student_name, "", "", "",
            "", "", "", "", "", session_number, login_time, ""
        ]
        
        print(f"Adding session row: {login_row}")
        sheet.append(login_row)
        workbook.save(file_path)
        workbook.close()
        print("Login session created successfully")
        
    except Exception as e:
        print(f"Error in save_login_to_excel: {e}")
        raise e

def student_login_view(request):
    """Student login page that saves data immediately upon login"""
    error_message = None
    success_message = None
    today_date = date.today().strftime("%Y-%m-%d")

    # Handle redirect to feedback form
    if request.method == 'GET' and 'redirect_to_feedback' in request.GET:
        if not request.session.get('student_name'):
            error_message = "❌ Please login first before accessing feedback form"
        else:
            return redirect('feedback_form')

    if request.method == 'POST':
        # Handle redirect button
        if 'redirect_to_feedback' in request.POST:
            if not request.session.get('student_name'):
                error_message = "❌ Please login first before accessing feedback form"
            else:
                return redirect('feedback_form')
            
        else:
            student_name = request.POST.get('student_name', '').strip()
            selected_date = request.POST.get('date', '').strip()
            state = request.POST.get('state', '').strip()
            
            if student_name and selected_date and state:
                login_time = datetime.now().strftime("%H:%M:%S")
                
                # Check if there's an existing incomplete session
                existing_session = check_existing_incomplete_session(student_name, selected_date, state)
                
                if existing_session:
                    # Update existing session with login time
                    row_num, session_number = existing_session
                    if update_existing_session_with_login(student_name, selected_date, state, login_time, row_num):
                        # Store student info in session
                        request.session['student_name'] = student_name
                        request.session['selected_date'] = selected_date
                        request.session['state'] = state
                        request.session['login_time'] = login_time
                        request.session['login_datetime'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        request.session['session_number'] = session_number
                        
                        success_message = f"✅ Login successful! Resumed Session #{session_number} for {student_name} in {state.title()}"
                    else:
                        error_message = "❌ Error updating existing session"
                else:
                    # Create new session
                    session_number = get_student_session_number(student_name, selected_date, state)
                    
                    # Store student info in session
                    request.session['student_name'] = student_name
                    request.session['selected_date'] = selected_date
                    request.session['state'] = state
                    request.session['login_time'] = login_time
                    request.session['login_datetime'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    request.session['session_number'] = session_number
                    
                    # Save login data to Excel
                    try:
                        save_login_to_excel(student_name, selected_date, state, login_time, session_number)
                        success_message = f"✅ Login successful! Session #{session_number} started for {student_name} in {state.title()}"
                    except Exception as e:
                        error_message = f"❌ Login successful but error saving data: {str(e)}"
                
            else:
                error_message = "❌ Please enter your name, select a date, and choose a state"

    return render(request, 'feedback/student_login.html', {
        'error_message': error_message,
        'success_message': success_message,
        'today_date': today_date,
        'show_feedback_redirect': request.session.get('student_name') is not None,
        'states': ['bangalore', 'hyderabad', 'odisha']
    })

def feedback_view(request):
    """Updated feedback view that updates the existing session row"""
    
    # Handle redirect to student login
    if request.method == 'GET' and 'redirect_to_login' in request.GET:
        return redirect('student_login')
    
    if request.method == 'POST' and 'redirect_to_login' in request.POST:
        return redirect('student_login')
    
    # Get data from session
    student_name = request.session.get('student_name')
    selected_date = request.session.get('selected_date')
    login_time = request.session.get('login_time')
    state = request.session.get('state')
    session_number = request.session.get('session_number')
    
    form = FeedbackForm()

    if request.method == 'POST':
        print("POST request received")
        print("POST data:", request.POST)
        
        # Extract data directly from request.POST
        feedback_data = {
            'state': state or request.POST.get('state', '').strip(),
            'student_name': request.POST.get('student_name', '').strip() or student_name,
            'trainer_name': request.POST.get('trainer_name', '').strip(),
            'course': request.POST.get('course', '').strip(),
            'slot_timings': request.POST.get('slot_timings', '').strip(),
            'understanding': request.POST.get('understanding', '').strip(),
            'engagement': request.POST.get('engagement', '').strip(),
            'overall': request.POST.get('overall', '').strip(),
            'homework': request.POST.get('homework', '').strip(),
            'parents_feedback': request.POST.get('parents_feedback', '').strip(),
        }
        
        print("Feedback data:", feedback_data)
        
        # Validate required fields
        required_fields = ['state', 'student_name', 'trainer_name', 'course', 'slot_timings', 
                          'understanding', 'engagement', 'overall']
        missing_fields = [field for field in required_fields if not feedback_data[field]]
        
        if not missing_fields:
            state_lower = feedback_data["state"].lower()
            file_path = get_excel_file(state_lower)
            
            # Use session number from session, or get/create a new one
            if not session_number:
                # Check if there's an existing session for this student/date first
                existing_session_row = find_existing_session_row(feedback_data["student_name"], 
                                                               selected_date or date.today().strftime("%Y-%m-%d"), 
                                                               state_lower)
                if existing_session_row:
                    session_number = existing_session_row[5] if len(existing_session_row) > 5 and existing_session_row[5] else 1
                else:
                    session_number = get_student_session_number(feedback_data["student_name"], 
                                                              selected_date or date.today().strftime("%Y-%m-%d"), 
                                                              state_lower)

            # Add session info
            feedback_data["selected_date"] = selected_date or date.today().strftime("%Y-%m-%d")
            feedback_data["login_time"] = login_time or ""
            feedback_data["logout_time"] = datetime.now().strftime("%H:%M:%S")

            try:
                # Try to update existing session, if not found, create new
                if session_number and update_existing_session(feedback_data, state_lower, session_number):
                    print("Updated existing session with feedback")
                else:
                    save_feedback_to_excel(feedback_data, state_lower, session_number or 1)
                    print("Created new session with feedback")
                
                # Clear session after successful submission
                request.session.flush()

                return render(request, 'feedback/feedback_form.html', {
                    'form': FeedbackForm(),
                    'student_name': feedback_data["student_name"],
                    'selected_date': feedback_data["selected_date"],
                    'success': True,
                    'session_number': session_number,
                    'show_redirect': True,
                    'show_login_redirect': True
                })
            except Exception as e:
                print(f"Error saving data: {e}")
                return render(request, 'feedback/feedback_form.html', {
                    'form': form,
                    'student_name': student_name,
                    'selected_date': selected_date,
                    'error': f"Error saving feedback: {str(e)}",
                    'show_login_redirect': True
                })
        else:
            print(f"Missing required fields: {missing_fields}")
            return render(request, 'feedback/feedback_form.html', {
                'form': form,
                'student_name': student_name,
                'selected_date': selected_date,
                'error': f"Please fill in all required fields: {', '.join(missing_fields)}",
                'show_login_redirect': True
            })

    return render(request, 'feedback/feedback_form.html', {
        'form': form, 
        'student_name': student_name,
        'selected_date': selected_date,
        'show_login_redirect': True
    })

def find_existing_session_row(student_name, selected_date, state):
    """Find existing session row for the student on the given date"""
    file_path = get_excel_file(state.lower())
    if not os.path.exists(file_path):
        return None
    
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if (row and len(row) > 1 and row[0] and row[1] and 
                str(row[0]) == selected_date and 
                row[1].strip().lower() == student_name.strip().lower()):
                workbook.close()
                return row
        
        workbook.close()
        return None
    except Exception as e:
        print(f"Error finding existing session: {e}")
        return None

def update_existing_session(feedback_data, state, session_number):
    """Try to update existing session with feedback data"""
    file_path = get_excel_file(state.lower())
    
    if not os.path.exists(file_path):
        return False

    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        
        # Find the session row to update
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if (row and len(row) > 10 and row[0] and row[1] and row[10] and
                str(row[0]) == feedback_data["selected_date"] and 
                row[1].strip().lower() == feedback_data["student_name"].strip().lower() and
                int(row[10]) == session_number):
                
                # Update the row with feedback data
                sheet.cell(row=row_num, column=3, value=feedback_data["trainer_name"])  # Trainer Name
                sheet.cell(row=row_num, column=4, value=feedback_data["course"])  # Course
                sheet.cell(row=row_num, column=5, value=feedback_data["slot_timings"])  # Slot Timings
                sheet.cell(row=row_num, column=6, value=feedback_data["understanding"])  # Understanding
                sheet.cell(row=row_num, column=7, value=feedback_data["engagement"])  # Engagement
                sheet.cell(row=row_num, column=8, value=feedback_data["overall"])  # Overall Feedback
                sheet.cell(row=row_num, column=9, value=feedback_data.get("homework", "No Homework"))  # Homework
                sheet.cell(row=row_num, column=10, value=feedback_data.get("parents_feedback", "No Parents Feedback"))  # Parents Feedback
                sheet.cell(row=row_num, column=13, value=feedback_data["logout_time"])  # Logout Time (column M)
                
                workbook.save(file_path)
                workbook.close()
                return True
        
        workbook.close()
        return False
        
    except Exception as e:
        print(f"Error updating existing session: {e}")
        return False

def save_feedback_to_excel(feedback_data, state, session_number):
    """Fallback: Create new feedback entry if session update fails"""
    file_path = get_excel_file(state)
    print(f"Saving to file: {file_path}")
    
    # Create file if it doesn't exist
    if not os.path.exists(file_path):
        print("Creating new Excel file")
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Feedback Data"
        sheet.append([
            "Date", "Student Name", "Trainer Name", "Course", "Slot Timings", "Understanding",
            "Engagement", "Overall Feedback", "Homework", "Parents Feedback", "Session Number", "Login Time", "Logout Time"
        ])
        workbook.save(file_path)
        workbook.close()

    # Open existing file and append data
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        
        new_row = [
            feedback_data.get("selected_date", date.today().strftime("%Y-%m-%d")),
            feedback_data.get("student_name", "N/A"),
            feedback_data.get("trainer_name", "N/A"),
            feedback_data.get("course", "N/A"),
            feedback_data.get("slot_timings", "N/A"),
            feedback_data.get("understanding", "N/A"),
            feedback_data.get("engagement", "N/A"),
            feedback_data.get("overall", "N/A"),
            feedback_data.get("homework", "No Homework"),
            feedback_data.get("parents_feedback", "No Parents Feedback"),
            session_number,
            "",  # Login time empty - to be filled by student login
            feedback_data.get("logout_time", "N/A")
        ]
        
        print(f"Adding row: {new_row}")
        sheet.append(new_row)
        workbook.save(file_path)
        workbook.close()
        print("File saved successfully")
        
    except Exception as e:
        print(f"Error in save_feedback_to_excel: {e}")
        raise e

def download_excel(request, state):
    if not request.session.get('is_admin'):
        return redirect('login')

    file_path = get_excel_file(state)
    if os.path.exists(file_path):
        with open(file_path, 'rb') as f:
            response = HttpResponse(
                f.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="feedback_{state}.xlsx"'
            return response

    return HttpResponse("❌ No feedback data found", status=404)

def login_view(request):
    request.session.flush()  # Clear old session
    error_message = None

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()

        if username == "user1" and password == "1234":
            request.session['is_student'] = True
            return redirect('student_login')
        elif username == "admin" and password == "4321":
            request.session['is_admin'] = True
            return redirect('admin_dashboard')
        else:
            error_message = "❌ Invalid username or password"

    return render(request, 'feedback/login.html', {'error_message': error_message})

def admin_dashboard(request):
    if not request.session.get('is_admin'):
        return redirect('login')

    feedback_files = [
        f for f in os.listdir(FEEDBACK_DIR) if f.endswith(".xlsx")
    ]
    feedback_data = [{"name": f[9:-5], "file": f} for f in feedback_files]

    return render(request, "feedback/admin_dashboard.html", {"feedback_data": feedback_data})

def logout_view(request):
    request.session.flush()
    return redirect('login')